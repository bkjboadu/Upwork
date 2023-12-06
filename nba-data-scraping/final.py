from nba_api.stats.endpoints import leagueleaders,playercareerstats
from nba_api.stats.static import players
import pandas as pd
import requests,openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pathlib import Path

Path('statistic').mkdir(exist_ok=True)

per_mode = 'Totals'
season_id = '2023-24'

HEADERS = {
    'Connection': 'keep-alive',
    'Accept': 'application/json, text/plain, */*',
    'x-nba-stats-token': 'true',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36',
    'x-nba-stats-origin': 'stats',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-Mode': 'cors',
    'Referer': 'https://stats.nba.com/',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-US,en;q=0.9',
}

save_column = ['Player','Team','Age','GP','PPG','FTA','FTM','FT%','3PA','3PM','3P%','FGA','FGM','FG%']
required_headers = ['PLAYER_NAME','TEAM_ABBREVIATION','AGE','GP','PTS','FTA','FTM','FT_PCT','FG3A','FG3M','FG3_PCT','FGA','FGM','FG_PCT']
active_players = players.get_active_players()

def extract(player_info_url, excel_file_name, per_mode='Totals', season_id='2023-24'):
    response = requests.get(url=player_info_url, headers=HEADERS)

    def create_excel_doc(response, excel_name):
        '''Creates a pandas dataframe from response and converts it to an excel file'''

        player_stats = response['resultSets'][0]
        excel_headers = player_stats['headers']
        datasets = player_stats['rowSet']

        dataframe = pd.DataFrame(datasets, columns=excel_headers)
        dataframe.to_excel(excel_name,index=False)
        return dataframe

    if response.status_code == 200:
        print('JSON file extracted successfully')
        response = response.json()
        print('Saving to XLSX file')
        dataframe = create_excel_doc(response, excel_file_name)
    else:
        print('Data was not extracted')
        print('Please check your internet connection')
    return dataframe

def get_career_stats():
    '''get player career statistics'''
    career_average_df = pd.DataFrame()
    for player in active_players:
        career_stats = playercareerstats.PlayerCareerStats(player_id=player['id'])
        data_frames = career_stats.get_data_frames()
        df = data_frames[0]
        df['EFG_PCT'] =  (df['FGM'] + (0.5 * df['FG3M'])) / df['FGA']
        columns = ['PLAYER_ID','GP', 'GS', 'MIN', 'FGM', 'FGA', 'FG_PCT', 'FG3M', 'FG3A',
       'FG3_PCT', 'FTM', 'FTA', 'FT_PCT', 'OREB', 'DREB', 'REB', 'AST', 'STL',
       'BLK', 'TOV', 'PF', 'PTS','EFG_PCT']
        df = df[columns]
        df = df.mean().to_frame().T
        df['PLAYER_NAME'] = player['full_name']

        career_average_df = pd.concat([career_average_df, df], ignore_index=True)
    career_average_df.to_excel('statistic/career_averages.xlsx',index=False)
    return career_average_df

def get_player_stats():
    player_info_url = (
        'https://stats.nba.com/stats/leaguedashplayerstats?'
        'College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick='
        '&DraftYear=&GameScope=&GameSegment=&Height=&LastNGames=0&LeagueID=00'
        '&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0'
        '&PaceAdjust=N&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&PlusMinus=N'
        '&Rank=N&Season={}&SeasonSegment=&SeasonType=Regular+Season&ShotClockRange='
        '&StarterBench=&TeamID=0&TwoWay=0&VsConference=&VsDivision=&Weight='
    ).format(per_mode, season_id)
    extract(player_info_url, 'statistic/playerstats.xlsx', per_mode, season_id)

def get_last10game_stats():
    last10_info_url = (
        'https://stats.nba.com/stats/leaguedashplayerstats?'
        'College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick='
        '&DraftYear=&GameScope=&GameSegment=&Height=&LastNGames=0&LeagueID=00'
        '&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0'
        '&PaceAdjust=N&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&PlusMinus=N'
        '&Rank=N&Season={}&SeasonSegment=&SeasonType=Regular+Season&ShotClockRange='
        '&StarterBench=&TeamID=0&TwoWay=0&VsConference=&VsDivision=&Weight='
    ).format(per_mode, season_id) 
    extract(last10_info_url,'statistic/last10games.xlsx', per_mode, season_id)


def get_clutch_stats():
    clutch_info_url = 'https://stats.nba.com/stats/leaguedashplayerclutch?AheadBehind=Ahead%20or%20Behind&ClutchTime=Last%205%20Minutes&College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick=&DraftYear=&GameScope=&GameSegment=&Height=&ISTRound=&LastNGames=0&LeagueID=00&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&PlusMinus=N&PointDiff=5&Rank=N&Season={}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&StarterBench=&TeamID=0&VsConference=&VsDivision=&Weight='.format(per_mode, season_id)
    extract(clutch_info_url,'statistic/playerclutch.xlsx', per_mode, season_id)

def get_range_stats():
    range_url = 'https://stats.nba.com/stats/leaguedashplayershotlocations?College=&Conference=&Country=&DateFrom=&DateTo=&DistanceRange=5ft%20Range&Division=&DraftPick=&DraftYear=&GameScope=&GameSegment=&Height=&ISTRound=&LastNGames=0&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&PlusMinus=N&Rank=N&Season={}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&StarterBench=&TeamID=0&VsConference=&VsDivision=&Weight='.format(per_mode,season_id)
    response = requests.get(url=range_url,headers=HEADERS)
    if response.status_code == 200:
        response = response.json()
    data = response['resultSets']['rowSet']
    data = [x[:2] + x[3:5] + x[6:24] for x in data]
    column1 = response['resultSets']['headers'][0]['columnNames'][:-3]
    column2 = ['FGM','FGA','FG_PCT',]
    labels = []
    for label1 in column1:
        for label2 in column2:
            final = label1 + '-' + label2
            labels.append(final)
    columns = ['PLAYER_ID','PLAYER_NAME','TEAM','AGE'] + labels
    assert len(data[0]) == len(columns)
    dataframe = pd.DataFrame(data,columns=columns)
    dataframe.to_excel('statistic/range_stats.xlsx',index=False)
    return dataframe

close_distance_0_2='https://stats.nba.com/stats/leaguedashplayerptshot?CloseDefDistRange=0-2%20Feet%20-%20Very%20Tight&College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick=&DraftYear=&DribbleRange=&GameScope=&GameSegment=&GeneralRange=&Height=&LastNGames=0&LeagueID=00&Location=&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&Season={}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&ShotDistRange=&StarterBench=&TeamID=0&TouchTimeRange=&VsConference=&VsDivision=&Weight='.format(per_mode,season_id)
close_distance_2_4='https://stats.nba.com/stats/leaguedashplayerptshot?CloseDefDistRange=2-4%20Feet%20-%20Tight&College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick=&DraftYear=&DribbleRange=&GameScope=&GameSegment=&GeneralRange=&Height=&LastNGames=0&LeagueID=00&Location=&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&Season={}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&ShotDistRange=&StarterBench=&TeamID=0&TouchTimeRange=&VsConference=&VsDivision=&Weight='.format(per_mode,season_id)
close_distance_4_6='https://stats.nba.com/stats/leaguedashplayerptshot?CloseDefDistRange=4-6%20Feet%20-%20Open&College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick=&DraftYear=&DribbleRange=&GameScope=&GameSegment=&GeneralRange=&Height=&LastNGames=0&LeagueID=00&Location=&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&Season={}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&ShotDistRange=&StarterBench=&TeamID=0&TouchTimeRange=&VsConference=&VsDivision=&Weight='.format(per_mode,season_id)
close_distance_6_='https://stats.nba.com/stats/leaguedashplayerptshot?CloseDefDistRange=6%2B%20Feet%20-%20Wide%20Open&College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick=&DraftYear=&DribbleRange=&GameScope=&GameSegment=&GeneralRange=&Height=&LastNGames=0&LeagueID=00&Location=&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PerMode={}&Period=0&PlayerExperience=&PlayerPosition=&Season={}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&ShotDistRange=&StarterBench=&TeamID=0&TouchTimeRange=&VsConference=&VsDivision=&Weight='.format(per_mode,season_id)


def get_closest_distance(url,save_to):
    extract(url,save_to, per_mode, season_id)
    
def extract_all():
    get_player_stats()
    get_last10game_stats()
    get_range_stats()
    get_clutch_stats()
    get_closest_distance(close_distance_0_2,'statistic/close_distance_0_2.xlsx')
    get_closest_distance(close_distance_2_4,'statistic/close_distance_2_4.xlsx')
    get_closest_distance(close_distance_4_6,'statistic/close_distance_4_6.xlsx')
    get_closest_distance(close_distance_6_,'statistic/close_distance_6_.xlsx')
    get_career_stats()

def process_traditional(traditional):
    save_column_as = ['PLAYER_ID','PLAYER_NAME','Team','Age','GP','PPG','EFG_%','FTA','FTM','FT%','3PA','3PM','3P%','FGA','FGM','FG%']
    required_headers = ['PLAYER_ID','PLAYER_NAME','TEAM_ABBREVIATION','AGE','GP','PTS','EFG_PCT','FTA','FTM','FT_PCT','FG3A','FG3M','FG3_PCT','FGA','FGM','FG_PCT']
    traditional['EFG_PCT'] =  (traditional['FGM'] + (0.5 * traditional['FG3M'])) / traditional['FGA']
    traditional = traditional[required_headers]
    traditional.columns = save_column_as
    return traditional

def process_last10(last10_games):
    save_column_as = ['PLAYER_ID','PLAYER_NAME','FG%','FG3%','FT%','EFG%']
    save_column_as = ['PLAYER_ID', 'PLAYER_NAME'] + ['last10'+x for x in save_column_as[2:]]
    required_headers = ['PLAYER_ID','PLAYER_NAME','FG_PCT','FG3_PCT','FT_PCT','EFG_PCT']
    last10_games['EFG_PCT'] =  (last10_games['FGM'] + (0.5 * last10_games['FG3M'])) / last10_games['FGA']
    last10_games = last10_games[required_headers]
    last10_games.columns = save_column_as
    return last10_games

def process_career(career):
    career_cols = ['PLAYER_ID','PLAYER_NAME','FG_PCT','FG3_PCT','FT_PCT','EFG_PCT']
    save_column_as = ['PLAYER_ID','PLAYER_NAME','FG%','FG3%','FT%','EFG%']
    save_column_as = ['PLAYER_ID', 'PLAYER_NAME'] + ['career'+x for x in save_column_as[2:]]
    career['EFG_PCT'] = (career['FGM'] + (0.5 * career['FG3M'])) / career['FGA']
    career = career[career_cols]
    career.columns = save_column_as
    return career

def process_clutch(clutch):
    save_column_as = ['PLAYER_ID','PLAYER_NAME','FGM','FGA','FG%','FG3M','FG3A','FG3%','FTM','FTA','FT%']
    save_column_as = ['PLAYER_ID', 'PLAYER_NAME'] + ['clutch'+x for x in save_column_as[2:]]
    required_headers = ['PLAYER_ID','PLAYER_NAME','FGM','FGA','FG_PCT','FG3M','FG3A','FG3_PCT','FTM','FTA','FT_PCT']
    clutch = clutch[required_headers]
    clutch.columns = save_column_as
    return clutch

def process_range(shoot):
    del shoot['TEAM']
    del shoot['AGE']
    return shoot

def process_closest_defender(cd_ft,ft):
    required_headers = ['PLAYER_ID', 'PLAYER_NAME','FGA_FREQUENCY','FGM', 'FGA', 'FG_PCT', 'EFG_PCT', 'FG2A_FREQUENCY', 'FG2M', 'FG2A','FG2_PCT', 'FG3A_FREQUENCY', 'FG3M', 'FG3A', 'FG3_PCT']
    save_column_as = ['PLAYER_ID', 'PLAYER_NAME','FGA_FREQUENCY','FGM', 'FGA', 'FG%', 'EFG%', 'FG2A_FREQUENCY', 'FG2M', 'FG2A','FG2%', 'FG3A_FREQUENCY', 'FG3M', 'FG3A', 'FG3%']
    cd_ft = cd_ft[required_headers]
    edited_save_column_as = ['PLAYER_ID', 'PLAYER_NAME'] + [str(ft)+x for x in save_column_as[2:]]
    cd_ft.columns = edited_save_column_as
    return cd_ft

def get_datasets():
    traditional = process_traditional(pd.read_excel('statistic/playerstats.xlsx'))
    last10_games = process_last10(pd.read_excel('statistic/last10games.xlsx'))
    clutch = process_clutch(pd.read_excel('statistic/playerclutch.xlsx'))
    shoot = process_range(pd.read_excel('statistic/range_stats.xlsx'))
    career = process_career(pd.read_excel('statistic/career_averages.xlsx'))
    cd_0_2 = process_closest_defender(pd.read_excel('statistic/close_distance_0_2.xlsx'),'0-2ft')
    cd_2_4 = process_closest_defender(pd.read_excel('statistic/close_distance_2_4.xlsx'),'2_4ft')
    cd_4_6 = process_closest_defender(pd.read_excel('statistic/close_distance_4_6.xlsx'),'4-6ft')
    cd_6_ = process_closest_defender(pd.read_excel('statistic/close_distance_6_.xlsx'),'6+ft')
    datasets = [traditional,career,last10_games,clutch,shoot,cd_0_2,cd_2_4,cd_4_6,cd_6_]
    return datasets

def merge_datasets(dataset):
    traditional,career,last10_games,clutch,shoot,cd_0_2,cd_2_4,cd_4_6,cd_6_ = dataset

    # Merge dataframes one by one on the common columns 'PLAYER_ID' and 'PLAYER_NAME'
    merged_df = traditional.merge(career, on=['PLAYER_ID', 'PLAYER_NAME'], how='left') \
                    .merge(last10_games, on=['PLAYER_ID', 'PLAYER_NAME'], how='left') \
                    .merge(clutch, on=['PLAYER_ID', 'PLAYER_NAME'], how='left') \
                    .merge(shoot, on=['PLAYER_ID', 'PLAYER_NAME'], how='left') \
                    .merge(cd_0_2, on=['PLAYER_ID', 'PLAYER_NAME'], how='left') \
                    .merge(cd_2_4, on=['PLAYER_ID', 'PLAYER_NAME'], how='left') \
                    .merge(cd_4_6, on=['PLAYER_ID', 'PLAYER_NAME'], how='left') \
                    .merge(cd_6_, on=['PLAYER_ID', 'PLAYER_NAME'], how='left')

    del merged_df['PLAYER_ID']
    merged_df.to_excel('all_player_stats.xlsx',index=False)
    print('Done')

def save_to_template():
    data_x = pd.read_excel('all_player_stats.xlsx')
    nba = openpyxl.load_workbook('nba_template.xlsx')
    nba_sheet = nba.active
    today_date = datetime.now().date()
    formatted_date = today_date.strftime("%Y-%m-%d")
    starting_row = 4

    for index,df_row in data_x.iterrows():
        for x,_ in enumerate(df_row):
            row = starting_row+index
            nba_sheet.cell(row=row,column=x+1).value = df_row[x]

    filename = f'stats.xlsx'
    nba.save(filename)
    print('Copied values into nba_template.xlsx')


if __name__ == '__main__':
    print('extracting datasets from nba site')
    extract_all()
    
    print('Performing data preprocessing for template')
    merge_datasets(get_datasets())

    save_to_template()




